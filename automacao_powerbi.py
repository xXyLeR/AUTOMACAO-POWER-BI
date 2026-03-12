"""
========================================================
  AUTOMAÇÃO POWER BI + EXCEL (Google Drive)
  Novidades:
    - Clique automático em "Atualizar" via pyautogui
    - Log visual em HTML com histórico de mudanças
========================================================
"""

import os
import sys
import json
import time
import logging
import hashlib
import subprocess
import shutil
from datetime import datetime
from pathlib import Path

# ── Dependências externas ──────────────────────────────────────────────────────
try:
    import openpyxl
    import schedule
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
except ImportError:
    print("[ERRO] Dependências não instaladas. Execute 'instalar_dependencias.bat' primeiro.")
    sys.exit(1)

try:
    import pyautogui
    import pygetwindow as gw
    PYAUTOGUI_DISPONIVEL = True
except ImportError:
    PYAUTOGUI_DISPONIVEL = False
    print("[AVISO] pyautogui não instalado — clique automático desativado.")


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURAÇÕES
# ══════════════════════════════════════════════════════════════════════════════

CONFIG_PATH = Path(__file__).parent / "config.json"

def carregar_config() -> dict:
    if not CONFIG_PATH.exists():
        print(f"[ERRO] config.json não encontrado em: {CONFIG_PATH}")
        sys.exit(1)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

CONFIG = carregar_config()


# ══════════════════════════════════════════════════════════════════════════════
#  LOGGING
# ══════════════════════════════════════════════════════════════════════════════

log_path = Path(__file__).parent / "logs" / f"automacao_{datetime.now():%Y%m%d}.log"
log_path.parent.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(log_path, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("AutoPBI")


# ══════════════════════════════════════════════════════════════════════════════
#  LOG VISUAL EM HTML
# ══════════════════════════════════════════════════════════════════════════════

HTML_LOG_PATH = Path(__file__).parent / "logs" / "historico.html"
_eventos_sessao: list[dict] = []

def _registrar_evento_html(tipo: str, titulo: str, detalhes: list[str], mudancas: dict = None):
    """Armazena um evento na memória para renderizar no HTML depois."""
    _eventos_sessao.append({
        "ts": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "tipo": tipo,          # "sucesso" | "aviso" | "erro" | "info"
        "titulo": titulo,
        "detalhes": detalhes,
        "mudancas": mudancas or {},
    })
    _salvar_html_log()


def _salvar_html_log():
    """Gera o arquivo HTML completo com todos os eventos da sessão + histórico."""
    # Carrega eventos anteriores salvo em JSON sidecar
    sidecar = HTML_LOG_PATH.with_suffix(".json")
    historico: list[dict] = []
    if sidecar.exists():
        try:
            with open(sidecar, "r", encoding="utf-8") as f:
                historico = json.load(f)
        except Exception:
            historico = []

    # Mescla: novos eventos + histórico anterior (mais recentes primeiro)
    todos = _eventos_sessao[::-1] + [e for e in historico if e not in _eventos_sessao]
    todos = todos[:200]  # Mantém até 200 eventos

    # Persiste histórico
    with open(sidecar, "w", encoding="utf-8") as f:
        json.dump(todos, f, ensure_ascii=False, indent=2)

    # Cores por tipo
    cores = {
        "sucesso": ("#d4edda", "#155724", "#28a745", "✅"),
        "aviso":   ("#fff3cd", "#856404", "#ffc107", "⚠️"),
        "erro":    ("#f8d7da", "#721c24", "#dc3545", "❌"),
        "info":    ("#d1ecf1", "#0c5460", "#17a2b8", "ℹ️"),
    }

    def card(ev: dict) -> str:
        bg, txt, borda, icone = cores.get(ev["tipo"], cores["info"])
        detalhes_html = ""
        if ev["detalhes"]:
            items = "".join(f"<li>{d}</li>" for d in ev["detalhes"])
            detalhes_html = f"<ul style='margin:6px 0 0 0;padding-left:18px;font-size:13px'>{items}</ul>"

        mudancas = ev.get("mudancas", {})
        mudancas_html = ""
        if mudancas.get("novas_abas") or mudancas.get("novas_colunas"):
            partes = []
            if mudancas.get("novas_abas"):
                partes.append(f"📋 Novas abas: <b>{', '.join(mudancas['novas_abas'])}</b>")
            for aba, cols in mudancas.get("novas_colunas", {}).items():
                partes.append(f"➕ [{aba}] Novas colunas: <b>{', '.join(cols)}</b>")
            for aba, cols in mudancas.get("colunas_removidas", {}).items():
                partes.append(f"➖ [{aba}] Colunas removidas: <b>{', '.join(cols)}</b>")
            if mudancas.get("abas_removidas"):
                partes.append(f"🗑️ Abas removidas: <b>{', '.join(mudancas['abas_removidas'])}</b>")
            items_m = "".join(f"<li>{p}</li>" for p in partes)
            mudancas_html = f"""
            <div style='margin-top:8px;padding:8px 12px;background:#fffff0;border-left:3px solid #ffc107;font-size:13px;border-radius:4px'>
              <b>Mudanças detectadas:</b>
              <ul style='margin:4px 0 0 0;padding-left:18px'>{items_m}</ul>
            </div>"""

        return f"""
        <div style='background:{bg};border-left:4px solid {borda};border-radius:6px;
                    padding:12px 16px;margin-bottom:10px;box-shadow:0 1px 3px rgba(0,0,0,.08)'>
          <div style='display:flex;justify-content:space-between;align-items:center'>
            <span style='font-size:15px;font-weight:600;color:{txt}'>{icone} {ev["titulo"]}</span>
            <span style='font-size:12px;color:#888;font-family:monospace'>{ev["ts"]}</span>
          </div>
          {detalhes_html}
          {mudancas_html}
        </div>"""

    cards_html = "\n".join(card(e) for e in todos)
    total_ok  = sum(1 for e in todos if e["tipo"] == "sucesso")
    total_err = sum(1 for e in todos if e["tipo"] == "erro")
    total_av  = sum(1 for e in todos if e["tipo"] == "aviso")

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
  <meta charset="UTF-8">
  <meta http-equiv="refresh" content="30">
  <title>Log — Automação Power BI</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f4f6fb; color: #222; }}
    header {{ background: linear-gradient(135deg,#1E4E79,#2e75b6); color: #fff; padding: 24px 32px; }}
    header h1 {{ font-size: 22px; font-weight: 700; }}
    header p  {{ font-size: 13px; opacity: .8; margin-top: 4px; }}
    .stats {{ display: flex; gap: 16px; padding: 20px 32px; flex-wrap: wrap; }}
    .stat {{ background: #fff; border-radius: 8px; padding: 14px 24px; flex: 1;
             box-shadow: 0 1px 4px rgba(0,0,0,.08); text-align: center; min-width: 120px; }}
    .stat .n {{ font-size: 28px; font-weight: 700; }}
    .stat .l {{ font-size: 12px; color: #888; margin-top: 2px; }}
    .container {{ padding: 0 32px 32px; max-width: 900px; margin: 0 auto; }}
    h2 {{ font-size: 15px; color: #555; margin: 0 0 14px; font-weight: 600; letter-spacing: .5px; text-transform: uppercase; }}
    .refresh {{ font-size: 11px; color: #aaa; text-align: right; margin-bottom: 10px; }}
    @media (max-width: 600px) {{ .stats {{ padding: 12px; }} .container {{ padding: 0 12px 24px; }} }}
  </style>
</head>
<body>
  <header>
    <h1>⚡ Automação Power BI + Excel</h1>
    <p>Log visual de atualizações • Atualiza a cada 30 segundos</p>
  </header>

  <div class="stats">
    <div class="stat"><div class="n" style="color:#28a745">{total_ok}</div><div class="l">Atualizações OK</div></div>
    <div class="stat"><div class="n" style="color:#ffc107">{total_av}</div><div class="l">Avisos</div></div>
    <div class="stat"><div class="n" style="color:#dc3545">{total_err}</div><div class="l">Erros</div></div>
    <div class="stat"><div class="n" style="color:#17a2b8">{len(todos)}</div><div class="l">Total de eventos</div></div>
  </div>

  <div class="container">
    <div class="refresh">Última geração: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}</div>
    <h2>Histórico de eventos</h2>
    {cards_html if cards_html else '<p style="color:#aaa;text-align:center;padding:40px">Nenhum evento registrado ainda.</p>'}
  </div>
</body>
</html>"""

    with open(HTML_LOG_PATH, "w", encoding="utf-8") as f:
        f.write(html)


# ══════════════════════════════════════════════════════════════════════════════
#  UTILITÁRIOS
# ══════════════════════════════════════════════════════════════════════════════

def hash_arquivo(caminho: str) -> str:
    h = hashlib.md5()
    with open(caminho, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def fazer_backup_excel(caminho_excel: str):
    if not CONFIG.get("fazer_backup", True):
        return
    backup_dir = Path(__file__).parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome = Path(caminho_excel).stem
    destino = backup_dir / f"{nome}_backup_{ts}.xlsx"
    shutil.copy2(caminho_excel, destino)
    log.info(f"Backup criado: {destino}")


def matar_processo_powerbi():
    try:
        subprocess.run(["taskkill", "/F", "/IM", "PBIDesktop.exe"],
                       capture_output=True, check=False)
        time.sleep(2)
        log.info("Power BI Desktop encerrado.")
    except Exception as e:
        log.warning(f"Não foi possível encerrar o Power BI: {e}")


def abrir_powerbi(caminho_pbix: str) -> bool:
    pbi_exe = CONFIG.get("powerbi_exe_path",
                         r"C:\Program Files\Microsoft Power BI Desktop\bin\PBIDesktop.exe")
    if not Path(pbi_exe).exists():
        log.error(f"Power BI Desktop não encontrado: {pbi_exe}")
        return False
    try:
        subprocess.Popen([pbi_exe, caminho_pbix])
        log.info(f"Power BI Desktop aberto: {caminho_pbix}")
        return True
    except Exception as e:
        log.error(f"Falha ao abrir Power BI: {e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  CLIQUE AUTOMÁTICO EM "ATUALIZAR" (pyautogui)
# ══════════════════════════════════════════════════════════════════════════════

def _aguardar_janela_pbi(timeout: int = 60) -> bool:
    """Espera o Power BI aparecer na barra de tarefas."""
    log.info(f"Aguardando janela do Power BI abrir (timeout: {timeout}s)...")
    inicio = time.time()
    while time.time() - inicio < timeout:
        janelas = gw.getWindowsWithTitle("Power BI Desktop")
        if janelas:
            log.info(f"Janela encontrada: '{janelas[0].title}'")
            return True
        time.sleep(2)
    log.warning("Timeout aguardando o Power BI abrir.")
    return False


def _aguardar_carregamento_pbi():
    """Aguarda o PBI carregar completamente (barra de loading some)."""
    segundos = CONFIG.get("aguardar_carregamento_pbi", 20)
    log.info(f"Aguardando {segundos}s para o Power BI carregar o modelo...")
    time.sleep(segundos)


def clicar_atualizar_pbi() -> bool:
    """
    Estratégia de clique automático em Atualizar:
      1. Foca a janela do Power BI
      2. Tenta localizar o botão "Atualizar" por imagem (se configurado)
      3. Fallback: usa atalho de teclado Alt+H, depois navega pela ribbon
    """
    if not PYAUTOGUI_DISPONIVEL:
        log.warning("pyautogui não disponível — clique automático ignorado.")
        return False

    cfg_click = CONFIG.get("auto_click", {})
    if not cfg_click.get("ativo", True):
        log.info("Clique automático desativado no config.json.")
        return False

    # Foca a janela
    janelas = gw.getWindowsWithTitle("Power BI Desktop")
    if not janelas:
        log.error("Janela do Power BI não encontrada para clicar em Atualizar.")
        return False

    janela = janelas[0]
    try:
        janela.activate()
        time.sleep(1)
    except Exception:
        pass

    pyautogui.FAILSAFE = True
    pyautogui.PAUSE = 0.3

    metodo = cfg_click.get("metodo", "teclado")  # "teclado" ou "imagem"

    if metodo == "imagem":
        # Tenta localizar o botão por imagem (requer screenshot salvo em auto_click.imagem_botao)
        img_path = cfg_click.get("imagem_botao", "")
        if img_path and Path(img_path).exists():
            try:
                local = pyautogui.locateCenterOnScreen(img_path, confidence=0.8)
                if local:
                    pyautogui.click(local)
                    log.info(f"✅ Botão 'Atualizar' clicado via imagem em {local}")
                    return True
                else:
                    log.warning("Botão não encontrado na tela por imagem. Usando atalho de teclado.")
            except Exception as e:
                log.warning(f"Falha ao localizar imagem: {e}. Usando atalho de teclado.")

    # Método padrão: atalho de teclado
    # Power BI: Home ribbon → botão Atualizar = Alt → H → R → A
    try:
        log.info("Enviando atalho de teclado para clicar em Atualizar (Alt → H → R → A)...")
        pyautogui.press("alt")
        time.sleep(0.5)
        pyautogui.press("h")       # Guia Página Inicial (Home)
        time.sleep(0.5)
        pyautogui.press("r")       # Atualizar (Refresh)
        time.sleep(0.5)
        # Aguarda o refresh completar
        aguardar_refresh = cfg_click.get("aguardar_refresh_segundos", 15)
        log.info(f"Aguardando {aguardar_refresh}s para o refresh concluir...")
        time.sleep(aguardar_refresh)
        log.info("✅ Atalho de Atualizar enviado ao Power BI.")
        return True
    except Exception as e:
        log.error(f"Falha ao enviar atalho de teclado: {e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  DETECÇÃO DE MUDANÇAS NO EXCEL
# ══════════════════════════════════════════════════════════════════════════════

ESTADO_ANTERIOR: dict = {}

def ler_estrutura_excel(caminho: str) -> dict:
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        estrutura = {}
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            headers = []
            for cell in next(ws.iter_rows(max_row=1), []):
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
            estrutura[nome_aba] = headers
        wb.close()
        return estrutura
    except Exception as e:
        log.error(f"Erro ao ler estrutura do Excel: {e}")
        return {}


def detectar_mudancas(antiga: dict, nova: dict) -> dict:
    mudancas = {
        "novas_abas": [],
        "abas_removidas": [],
        "novas_colunas": {},
        "colunas_removidas": {},
    }
    for aba in nova:
        if aba not in antiga:
            mudancas["novas_abas"].append(aba)
    for aba in antiga:
        if aba not in nova:
            mudancas["abas_removidas"].append(aba)
    for aba in nova:
        if aba in antiga:
            adicionadas = set(nova[aba]) - set(antiga[aba])
            removidas   = set(antiga[aba]) - set(nova[aba])
            if adicionadas:
                mudancas["novas_colunas"][aba] = list(adicionadas)
            if removidas:
                mudancas["colunas_removidas"][aba] = list(removidas)
    return mudancas


def ha_mudancas(m: dict) -> bool:
    return any([m["novas_abas"], m["abas_removidas"], m["novas_colunas"], m["colunas_removidas"]])


def resumo_mudancas(m: dict) -> list[str]:
    linhas = []
    if m["novas_abas"]:
        linhas.append(f"Novas abas: {', '.join(m['novas_abas'])}")
    if m["abas_removidas"]:
        linhas.append(f"Abas removidas: {', '.join(m['abas_removidas'])}")
    for aba, cols in m["novas_colunas"].items():
        linhas.append(f"[{aba}] Novas colunas: {', '.join(cols)}")
    for aba, cols in m["colunas_removidas"].items():
        linhas.append(f"[{aba}] Colunas removidas: {', '.join(cols)}")
    return linhas


# ══════════════════════════════════════════════════════════════════════════════
#  PIPELINE PRINCIPAL DE ATUALIZAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

def atualizar_powerbi(caminho_excel: str, mudancas: dict = None):
    mudancas = mudancas or {}
    log.info("═" * 60)
    log.info("🔄 INICIANDO ATUALIZAÇÃO DO POWER BI...")

    caminho_pbix = CONFIG.get("powerbi_pbix_path", "")
    if not caminho_pbix or not Path(caminho_pbix).exists():
        msg = f"Arquivo .pbix não encontrado: {caminho_pbix}"
        log.error(msg)
        _registrar_evento_html("erro", "Power BI não encontrado", [msg], mudancas)
        return False

    # Copia Excel para pasta local se configurado
    excel_local = CONFIG.get("excel_local_path", "")
    if excel_local and excel_local != caminho_excel:
        try:
            shutil.copy2(caminho_excel, excel_local)
            log.info(f"Excel copiado para: {excel_local}")
        except Exception as e:
            msg = f"Falha ao copiar Excel: {e}"
            log.error(msg)
            _registrar_evento_html("erro", "Falha ao copiar Excel", [msg], mudancas)
            return False

    # Fecha Power BI
    if CONFIG.get("fechar_powerbi_antes", True):
        matar_processo_powerbi()
        time.sleep(CONFIG.get("aguardar_segundos_antes_abrir", 3))

    # Abre Power BI
    sucesso_abrir = abrir_powerbi(caminho_pbix)
    if not sucesso_abrir:
        _registrar_evento_html("erro", "Falha ao abrir Power BI", ["Verifique o caminho no config.json"], mudancas)
        return False

    # Aguarda carregar e clica em Atualizar
    _aguardar_janela_pbi(timeout=90)
    _aguardar_carregamento_pbi()

    sucesso_click = clicar_atualizar_pbi()

    # Monta log HTML
    detalhes = []
    if sucesso_click:
        detalhes.append("Power BI aberto e botão Atualizar acionado automaticamente.")
    else:
        detalhes.append("Power BI aberto — clique automático não disponível (clique em Atualizar manualmente).")

    arquivo_excel = Path(caminho_excel).name
    detalhes.append(f"Arquivo Excel: {arquivo_excel}")

    if ha_mudancas(mudancas):
        detalhes += resumo_mudancas(mudancas)
        tipo_evento = "aviso"
        titulo_evento = "Atualização concluída com mudanças estruturais"
    else:
        tipo_evento = "sucesso"
        titulo_evento = "Atualização concluída com sucesso"

    _registrar_evento_html(tipo_evento, titulo_evento, detalhes, mudancas)

    log.info(f"✅ ATUALIZAÇÃO CONCLUÍDA! Log HTML: {HTML_LOG_PATH}")
    log.info("═" * 60)
    return True


# ══════════════════════════════════════════════════════════════════════════════
#  MONITORAMENTO (Watchdog)
# ══════════════════════════════════════════════════════════════════════════════

class MonitorExcel(FileSystemEventHandler):
    def __init__(self, caminho_excel: str):
        self.caminho_excel = os.path.abspath(caminho_excel)
        self.ultimo_hash   = hash_arquivo(self.caminho_excel)
        self.ultima_exec   = 0
        self.cooldown      = CONFIG.get("cooldown_segundos", 30)

    def on_modified(self, event):
        if event.is_directory:
            return
        if os.path.abspath(event.src_path) != self.caminho_excel:
            return

        agora = time.time()
        if agora - self.ultima_exec < self.cooldown:
            return

        time.sleep(2)
        try:
            novo_hash = hash_arquivo(self.caminho_excel)
        except Exception:
            return

        if novo_hash == self.ultimo_hash:
            return

        log.info(f"📝 Mudança detectada: {self.caminho_excel}")
        self.ultimo_hash = novo_hash
        self.ultima_exec = agora

        global ESTADO_ANTERIOR
        nova_estrutura = ler_estrutura_excel(self.caminho_excel)
        mudancas       = detectar_mudancas(ESTADO_ANTERIOR, nova_estrutura)
        ESTADO_ANTERIOR = nova_estrutura

        fazer_backup_excel(self.caminho_excel)
        atualizar_powerbi(self.caminho_excel, mudancas)


def iniciar_monitoramento():
    caminho_excel = CONFIG.get("excel_path", "")
    if not caminho_excel or not Path(caminho_excel).exists():
        log.error(f"Excel não encontrado: {caminho_excel}")
        sys.exit(1)

    global ESTADO_ANTERIOR
    ESTADO_ANTERIOR = ler_estrutura_excel(caminho_excel)
    log.info(f"📊 Estrutura inicial: {list(ESTADO_ANTERIOR.keys())}")

    _registrar_evento_html("info", "Monitoramento iniciado",
                           [f"Arquivo: {Path(caminho_excel).name}",
                            f"Log HTML: {HTML_LOG_PATH}"])

    pasta   = str(Path(caminho_excel).parent)
    handler = MonitorExcel(caminho_excel)
    obs     = Observer()
    obs.schedule(handler, pasta, recursive=False)
    obs.start()

    log.info("═" * 60)
    log.info("👁️  MONITORAMENTO ATIVO")
    log.info(f"   Arquivo  : {caminho_excel}")
    log.info(f"   Log HTML : {HTML_LOG_PATH}")
    log.info(f"   Cooldown : {CONFIG.get('cooldown_segundos', 30)}s")
    log.info("   Ctrl+C para encerrar.")
    log.info("═" * 60)
    return obs


# ══════════════════════════════════════════════════════════════════════════════
#  AGENDAMENTO PERIÓDICO
# ══════════════════════════════════════════════════════════════════════════════

def configurar_agendamento():
    agenda = CONFIG.get("agendamento", {})
    if not agenda.get("ativo", False):
        return

    intervalo  = agenda.get("intervalo_minutos", 60)
    horarios   = agenda.get("horarios_fixos", [])
    caminho    = CONFIG.get("excel_path", "")

    def executar():
        log.info("⏰ Execução agendada iniciada...")
        nova = ler_estrutura_excel(caminho)
        m    = detectar_mudancas(ESTADO_ANTERIOR, nova)
        atualizar_powerbi(caminho, m)

    if horarios:
        for h in horarios:
            schedule.every().day.at(h).do(executar)
            log.info(f"📅 Agendado às {h}")
    else:
        schedule.every(intervalo).minutes.do(executar)
        log.info(f"📅 Agendado a cada {intervalo} min")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    log.info("🚀 Automação Power BI v2.0 iniciada")

    modo = CONFIG.get("modo", "ambos")

    if modo == "unico":
        caminho = CONFIG.get("excel_path", "")
        atualizar_powerbi(caminho, {})
        return

    obs = None
    if modo in ("monitoramento", "ambos"):
        obs = iniciar_monitoramento()

    if modo in ("agendado", "ambos"):
        configurar_agendamento()

    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        log.info("⛔ Encerrado pelo usuário.")
        _registrar_evento_html("info", "Monitoramento encerrado", ["Encerrado manualmente pelo usuário."])
        if obs:
            obs.stop()
            obs.join()


if __name__ == "__main__":
    main()
